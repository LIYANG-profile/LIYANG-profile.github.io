"""Generate bilibili-ops-tracker.xlsx for portfolio download + repo archive."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = Path(__file__).resolve().parent
PUBLIC = ROOT / "portfolio" / "public" / "downloads"

VIEW, LIKE, FAV, COIN, SHARE, REPLY, DANMAKU = 3375, 1358, 182, 22, 15, 26, 18
BLUE = Font(name="Arial", color="0000FF")
BLACK = Font(name="Arial", color="000000")
GREEN = Font(name="Arial", color="008000")
HEADER = Font(name="Arial", bold=True, color="111111")
MUTED = Font(name="Arial", color="666666", size=9)
YELLOW = PatternFill("solid", fgColor="FFFF00")
HEADER_FILL = PatternFill("solid", fgColor="F3F0EA")
ACCENT_FILL = PatternFill("solid", fgColor="F8E6E2")
LINE = Border(
    left=Side(style="thin", color="D9D6CF"),
    right=Side(style="thin", color="D9D6CF"),
    top=Side(style="thin", color="D9D6CF"),
    bottom=Side(style="thin", color="D9D6CF"),
)


def style_header_row(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.border = LINE


def main() -> None:
    PUBLIC.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "00_说明"
    ws["A1"] = "B站账号运营追踪表（作品集展示）"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    meta = [
        ("账号", "不是电影乐队"),
        ("主页", "https://space.bilibili.com/505201963"),
        ("样本稿件", "BV1sVAKzUEy9"),
        ("快照日期", "2026-07-22"),
    ]
    for idx, (k, v) in enumerate(meta, 3):
        ws.cell(idx, 1, k).font = MUTED
        ws.cell(idx, 2, v).font = BLACK
    ws["A8"] = "字段颜色"
    ws["A9"] = "蓝色数字 = 手工/接口输入（可改）"
    ws["A10"] = "黑色数字 = 公式计算结果（勿直接改）"
    ws["A11"] = "绿色文字 = 引用其他工作表"
    ws["A12"] = "黄色底 = 待创作中心导出后补齐的私域指标"
    ws["A14"] = "工具链"
    ws["A15"] = "Python 拉取公开指标 → 本表公式复盘 → SQL 题材汇总 → Power BI 看板"
    ws["A17"] = "更新方式"
    ws["A18"] = "运行 etl_public_stats.py 刷新公开指标；完播/来源请粘贴创作中心导出到黄底列。"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 56

    raw = wb.create_sheet("01_原始数据")
    headers = [
        "bvid",
        "title",
        "category",
        "pubdate",
        "duration_sec",
        "views",
        "likes",
        "favorites",
        "coins",
        "shares",
        "replies",
        "danmaku",
        "avg_watch_sec",
        "completion_rate",
        "fan_delta",
        "source_top",
        "note",
    ]
    for i, h in enumerate(headers, 1):
        raw.cell(1, i, h)
    style_header_row(raw, 1, len(headers))

    values = [
        "BV1sVAKzUEy9",
        "【挽救计划】披着太空冒险外皮的童话故事，无剧透简评高司令新片",
        "电影前瞻",
        "2026-03-22",
        322,
        VIEW,
        LIKE,
        FAV,
        COIN,
        SHARE,
        REPLY,
        DANMAKU,
        None,
        None,
        None,
        None,
        "公开接口快照；黄底列待创作中心补齐",
    ]
    for i, v in enumerate(values, 1):
        cell = raw.cell(2, i, v)
        cell.border = LINE
        if 6 <= i <= 12:
            cell.font = BLUE
        elif 13 <= i <= 16:
            cell.font = BLUE
            cell.fill = YELLOW
        else:
            cell.font = BLACK

    for r in range(3, 8):
        for c in range(1, len(headers) + 1):
            cell = raw.cell(r, c, None)
            cell.border = LINE
            if 6 <= c <= 16:
                cell.font = BLUE
            if 13 <= c <= 16:
                cell.fill = YELLOW

    widths = [14, 48, 12, 12, 12, 10, 10, 10, 10, 10, 10, 10, 14, 14, 12, 16, 36]
    for i, w in enumerate(widths, 1):
        raw.column_dimensions[get_column_letter(i)].width = w

    calc = wb.create_sheet("02_指标计算")
    calc["A1"] = "稿件指标计算（引用 01_原始数据）"
    calc["A1"].font = Font(name="Arial", bold=True, size=12)
    calc_headers = [
        "bvid",
        "title",
        "category",
        "views",
        "likes",
        "favorites",
        "coins",
        "shares",
        "like_rate",
        "favorite_rate",
        "coin_rate",
        "share_rate",
        "interaction_rate",
        "insight",
    ]
    for i, h in enumerate(calc_headers, 1):
        calc.cell(3, i, h)
    style_header_row(calc, 3, len(calc_headers))

    for r in range(4, 9):
        src = r - 2
        refs = [
            f"='01_原始数据'!A{src}",
            f"='01_原始数据'!B{src}",
            f"='01_原始数据'!C{src}",
            f"='01_原始数据'!F{src}",
            f"='01_原始数据'!G{src}",
            f"='01_原始数据'!H{src}",
            f"='01_原始数据'!I{src}",
            f"='01_原始数据'!J{src}",
        ]
        for c, formula in enumerate(refs, 1):
            cell = calc.cell(r, c, formula)
            cell.font = GREEN
            cell.border = LINE

        rate_formulas = [
            f'=IF(OR(D{r}="",D{r}=0),"",E{r}/D{r})',
            f'=IF(OR(D{r}="",D{r}=0),"",F{r}/D{r})',
            f'=IF(OR(D{r}="",D{r}=0),"",G{r}/D{r})',
            f'=IF(OR(D{r}="",D{r}=0),"",H{r}/D{r})',
            f"=IF(OR(D{r}=\"\",D{r}=0),\"\",SUM('01_原始数据'!G{src}:L{src})/D{r})",
        ]
        for offset, formula in enumerate(rate_formulas):
            cell = calc.cell(r, 9 + offset, formula)
            cell.font = BLACK
            cell.number_format = "0.0%"
            cell.border = LINE

        insight = calc.cell(
            r,
            14,
            f'=IF(D{r}="","",IF(I{r}>=0.2,"高共鸣，检查投币/分享转化","观察播放与封面点击"))',
        )
        insight.font = BLACK
        insight.border = LINE

    calc["A10"] = "结论（样本稿）"
    calc["A10"].font = HEADER
    calc["A11"] = "点赞率约 40%：内容共鸣强；投币/分享偏低：片尾 CTA 与话题锚点可加强。"
    calc.merge_cells("A11:N11")
    calc["A11"].fill = ACCENT_FILL

    for i, w in enumerate([14, 42, 12, 10, 10, 10, 10, 10, 12, 12, 12, 12, 14, 40], 1):
        calc.column_dimensions[get_column_letter(i)].width = w

    topic = wb.create_sheet("03_题材对照")
    for i, h in enumerate(["题材", "样本", "状态", "播放", "点赞率", "运营结论"], 1):
        topic.cell(1, i, h)
    style_header_row(topic, 1, 6)
    topic_rows = [
        ["电影前瞻", "挽救计划", "已接入", VIEW, LIKE / VIEW, "高点赞、可沉淀影评模板"],
        ["财经口播", "港深通勤等", "待接入", None, None, "小红书侧作品，待统一口径"],
        ["微电影 / AI", "现金之城等", "待接入", None, None, "非本账号公开稿，作对照空位"],
    ]
    for r, row in enumerate(topic_rows, 2):
        for c, v in enumerate(row, 1):
            cell = topic.cell(r, c, v)
            cell.border = LINE
            cell.font = BLUE if c in (4, 5) and v is not None else BLACK
            if c == 5 and isinstance(v, float):
                cell.number_format = "0.0%"
    for i, w in enumerate([14, 16, 10, 10, 10, 36], 1):
        topic.column_dimensions[get_column_letter(i)].width = w

    funnel = wb.create_sheet("04_漏斗复盘")
    funnel["A1"] = "创作中心漏斗（待私域数据）"
    funnel["A1"].font = Font(name="Arial", bold=True, size=12)
    for i, h in enumerate(["阶段", "指标", "数值", "转化率", "备注"], 1):
        funnel.cell(3, i, h)
    style_header_row(funnel, 3, 5)
    stages = [
        ("曝光", "覆盖人数", None, "创作中心-流量来源"),
        ("点击", "播放", VIEW, "公开接口已填播放"),
        ("完播", "完播率/均播时长", None, "黄底待导出"),
        ("互动", "点赞+收藏+投币+分享+评+弹", LIKE + FAV + COIN + SHARE + REPLY + DANMAKU, "公开接口汇总"),
        ("关注", "涨粉", None, "黄底待导出"),
    ]
    for r, (stage, metric, value, note) in enumerate(stages, 4):
        funnel.cell(r, 1, stage).font = BLACK
        funnel.cell(r, 2, metric).font = BLACK
        value_cell = funnel.cell(r, 3, value)
        value_cell.font = BLUE
        if value is None:
            value_cell.fill = YELLOW
        funnel.cell(r, 5, note).font = MUTED
        for c in range(1, 6):
            funnel.cell(r, c).border = LINE

    funnel["D5"] = '=IF(OR(C4="",C4=0),"",C5/C4)'
    funnel["D5"].number_format = "0.0%"
    funnel["D5"].font = BLACK
    funnel["D6"] = "待补完播"
    funnel["D7"] = '=IF(OR(C5="",C5=0),"",C7/C5)'
    funnel["D7"].number_format = "0.0%"
    funnel["D7"].font = BLACK
    funnel["D8"] = "待补涨粉"
    for i, w in enumerate([10, 28, 14, 12, 28], 1):
        funnel.column_dimensions[get_column_letter(i)].width = w

    dash = wb.create_sheet("05_周报看板")
    dash["A1"] = "本周看板 · 样本稿"
    dash["A1"].font = Font(name="Arial", bold=True, size=14)
    dash["A3"] = "KPI"
    dash["A3"].font = HEADER
    kpis = [
        ("播放", VIEW, "0"),
        ("点赞", LIKE, "0"),
        ("点赞率", LIKE / VIEW, "0.0%"),
        ("互动率", (LIKE + FAV + COIN + SHARE + REPLY + DANMAKU) / VIEW, "0.0%"),
    ]
    for i, (label, val, fmt) in enumerate(kpis, 4):
        dash.cell(i, 1, label).font = MUTED
        cell = dash.cell(i, 2, val)
        cell.font = BLACK
        cell.number_format = fmt
        cell.fill = ACCENT_FILL

    dash["A9"] = "互动构成"
    dash["A9"].font = HEADER
    dash["A10"] = "类型"
    dash["B10"] = "数量"
    dash["A10"].font = HEADER
    dash["B10"].font = HEADER
    for i, (label, val) in enumerate(
        [("点赞", LIKE), ("收藏", FAV), ("评论", REPLY), ("投币", COIN), ("弹幕", DANMAKU), ("分享", SHARE)],
        11,
    ):
        dash.cell(i, 1, label).font = BLUE
        dash.cell(i, 2, val).font = BLUE

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Interaction Mix"
    data = Reference(dash, min_col=2, min_row=10, max_row=16)
    cats = Reference(dash, min_col=1, min_row=11, max_row=16)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 14
    chart.height = 8
    dash.add_chart(chart, "D9")

    dash["A18"] = "下周行动"
    dash["A18"].font = HEADER
    dash["A19"] = "1. 片尾增加关注/投币口播与字幕 CTA"
    dash["A20"] = "2. 标题/封面 A/B：强化「无剧透 + 情绪钩子」"
    dash["A21"] = "3. 接入创作中心完播曲线，定位流失段落"
    dash["A23"] = "Power BI 建议页"
    dash["A23"].font = HEADER
    dash["A24"] = "页1 总览 KPI｜页2 单稿互动构成｜页3 题材矩阵（见 powerbi_model.md）"
    dash.column_dimensions["A"].width = 28
    dash.column_dimensions["B"].width = 14

    out1 = OUT_DIR / "bilibili-ops-tracker.xlsx"
    out2 = PUBLIC / "bilibili-ops-tracker.xlsx"
    wb.save(out1)
    wb.save(out2)
    print(f"saved {out1}")
    print(f"saved {out2}")


if __name__ == "__main__":
    main()
